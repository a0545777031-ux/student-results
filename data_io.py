# -*- coding: utf-8 -*-
"""Excel input/output and PDF report generation for parsed student results."""
import os, io, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4, landscape
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import mm
import ar_shape as S

FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
_FONTS_READY = False
def _ensure_fonts():
    global _FONTS_READY
    if _FONTS_READY:
        return
    reg = os.path.join(FONT_DIR, "DejaVuSans.ttf")
    bold = os.path.join(FONT_DIR, "DejaVuSans-Bold.ttf")
    for cand in ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",):
        if not os.path.exists(reg) and os.path.exists(cand):
            reg = cand
    pdfmetrics.registerFont(TTFont("Ar", reg))
    pdfmetrics.registerFont(TTFont("Ar-Bold", bold if os.path.exists(bold) else reg))
    _FONTS_READY = True

COMP_LABELS = {"short_tests": "اختبارات قصيرة", "assessment": "أدوات تقييم",
               "final_exam": "نهاية الفصل", "total": "المجموع"}
TERM_LABELS = {"t1": "الفصل الأول", "t2": "الفصل الثاني"}
COMP_ORDER = ["short_tests", "assessment", "final_exam", "total"]

# ---------------- Excel export ----------------
def to_excel(data, out_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "النتائج"
    ws.sheet_view.rightToLeft = True
    hdr_fill = PatternFill("solid", fgColor="0E5A4D")
    hdr_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ["م", "اسم الطالب", "رقم الهوية", "المادة", "الفصل",
               "اختبارات قصيرة", "أدوات تقييم", "نهاية الفصل", "المجموع"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
    r = 2
    for st in data["students"]:
        for subj in data["subjects"]:
            comps = st["grades"].get(subj, {})
            if not comps:
                continue
            for term in ("t1", "t2"):
                if not any(term in comps.get(ck, {}) for ck in COMP_ORDER):
                    continue
                row = [st["seq"], st["name"], st["id"], subj, TERM_LABELS[term]]
                for ck in COMP_ORDER:
                    row.append(comps.get(ck, {}).get(term, ""))
                ws.append(row)
                for c in ws[r]:
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                r += 1
    widths = [5, 32, 14, 26, 14, 14, 12, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # summary sheet: subject averages (of total, t1)
    ws2 = wb.create_sheet("تحليل المواد")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["المادة", "متوسط المجموع (ف1)", "أعلى", "أدنى", "عدد الطلاب"])
    for c in ws2[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.border = border
        c.alignment = Alignment(horizontal="center")
    for subj in data["subjects"]:
        vals = []
        for st in data["students"]:
            v = st["grades"].get(subj, {}).get("total", {}).get("t1")
            if isinstance(v, (int, float)):
                vals.append(v)
        if vals:
            ws2.append([subj, round(sum(vals)/len(vals), 2), max(vals), min(vals), len(vals)])
    for i, w in enumerate([26, 18, 10, 10, 12], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    wb.save(out_path)
    return out_path

# ---------------- Excel import ----------------
def parse_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in rows if r and any(c is not None for c in r)]
    if not rows:
        return {"meta": {}, "subjects": [], "students": [], "components": []}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    # long format produced by our export
    if "اسم الطالب" in idx or "الاسم" in idx:
        name_i = idx.get("اسم الطالب", idx.get("الاسم"))
        id_i = idx.get("رقم الهوية")
        subj_i = idx.get("المادة")
        term_i = idx.get("الفصل")
        comp_cols = {"short_tests": idx.get("اختبارات قصيرة"), "assessment": idx.get("أدوات تقييم"),
                     "final_exam": idx.get("نهاية الفصل"), "total": idx.get("المجموع")}
        students = {}
        subjects = []
        for r in rows[1:]:
            nm = r[name_i] if name_i is not None and name_i < len(r) else None
            if not nm:
                continue
            nm = str(nm).strip()
            sid = str(r[id_i]).strip() if id_i is not None and id_i < len(r) and r[id_i] else ""
            subj = str(r[subj_i]).strip() if subj_i is not None and subj_i < len(r) and r[subj_i] else "عام"
            term = "t2" if (term_i is not None and term_i < len(r) and r[term_i] and "ثاني" in str(r[term_i])) else "t1"
            if subj not in subjects:
                subjects.append(subj)
            st = students.setdefault(nm, {"name": nm, "id": sid, "seq": len(students)+1, "grades": {}})
            for ck, ci in comp_cols.items():
                if ci is not None and ci < len(r) and isinstance(r[ci], (int, float)):
                    st["grades"].setdefault(subj, {}).setdefault(ck, {})[term] = float(r[ci])
        studs = list(students.values())
    else:
        # generic wide format: first col = name, remaining numeric cols = subjects
        subjects = [h for h in header[1:] if h]
        studs = []
        for i, r in enumerate(rows[1:], 1):
            nm = str(r[0]).strip() if r and r[0] else f"طالب {i}"
            st = {"name": nm, "id": "", "seq": i, "grades": {}}
            for j, subj in enumerate(subjects, 1):
                if j < len(r) and isinstance(r[j], (int, float)):
                    st["grades"].setdefault(subj, {}).setdefault("total", {})["t1"] = float(r[j])
            studs.append(st)
    from parser_core import _present_components
    return {"meta": {"school": "", "grade_class": ws.title, "year": ""},
            "subjects": subjects, "students": studs,
            "components": _present_components(studs)}

# ---------------- PDF report ----------------
def to_pdf(data, out_path, display_name="النتائج", term="t1", component="total"):
    _ensure_fonts()
    W, H = landscape(A4)
    c = canvas.Canvas(out_path, pagesize=(W, H))
    def rtl(x): return S.shape(str(x))
    subjects = data["subjects"]
    students = data["students"]
    meta = data.get("meta", {})

    def header():
        c.setFillColorRGB(0.055, 0.353, 0.302)
        c.rect(0, H-70, W, 70, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Ar-Bold", 18)
        c.drawRightString(W-30, H-38, rtl("تحليل نتائج الطلاب"))
        c.setFont("Ar", 11)
        sub = " | ".join([x for x in [meta.get("school",""), meta.get("grade_class",""), meta.get("year","")] if x])
        c.drawRightString(W-30, H-58, rtl(sub))
        c.setFillColorRGB(0.78, 0.64, 0.29)
        c.setFont("Ar-Bold", 12)
        c.drawString(30, H-40, rtl(display_name))

    header()
    y = H-95
    # table: rows=students, cols=subjects (values = chosen component/term)
    col_name_w = 150
    avail = W-60-col_name_w
    cw = avail/max(len(subjects), 1)
    x0 = W-30
    # header row
    c.setFillColorRGB(0.90, 0.94, 0.92)
    c.rect(30, y-6, W-60, 24, fill=1, stroke=0)
    c.setFillColorRGB(0.06, 0.30, 0.26)
    c.setFont("Ar-Bold", 9)
    c.drawRightString(x0, y, rtl("الطالب"))
    c.setFont("Ar-Bold", 7)
    maxc = max(4, int(cw/4.6))
    xx = x0-col_name_w+cw/2
    for subj in subjects:
        label = subj if len(subj) <= maxc else subj[:maxc-1]+"…"
        c.drawCentredString(xx, y, rtl(label)); xx -= cw
    y -= 24
    c.setFont("Ar", 9)
    lbl = f"{COMP_LABELS.get(component,'')} - {TERM_LABELS.get(term,'')}"
    for i, st in enumerate(students):
        if y < 40:
            c.setFont("Ar", 8); c.setFillColorRGB(0.5,0.5,0.5)
            c.drawCentredString(W/2, 20, rtl(lbl)); c.showPage()
            header(); y = H-95; c.setFont("Ar", 9)
        if i % 2 == 0:
            c.setFillColorRGB(0.97, 0.98, 0.97); c.rect(30, y-4, W-60, 16, fill=1, stroke=0)
        c.setFillColorRGB(0.1, 0.1, 0.1)
        nm = st["name"] if len(st["name"]) <= 24 else st["name"][:23]+"…"
        c.drawRightString(x0, y, rtl(nm))
        xx = x0-col_name_w+cw/2
        for subj in subjects:
            v = st["grades"].get(subj, {}).get(component, {}).get(term, "")
            c.drawCentredString(xx, y, "" if v == "" else f"{v:g}"); xx -= cw
        y -= 16
    c.setFont("Ar", 8); c.setFillColorRGB(0.5,0.5,0.5)
    c.drawCentredString(W/2, 20, rtl(lbl + "  —  " + display_name))
    c.save()
    return out_path
